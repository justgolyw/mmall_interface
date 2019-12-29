#!/usr/bin/env python
# -*- coding:utf-8 -*-

import openpyxl
import os, sys

base_path = os.getcwd()
sys.path.append(base_path)

class HandleExcel:
    # 加载excel
    def load_excel(self):
        file_path = os.path.abspath(os.path.join(base_path, "..")) + "/Case/mmall testcase.xlsx"
        wb = openpyxl.load_workbook(file_path)
        # sheet = wb.active
        return wb

    # 获取行数
    def get_rows(self):
        wb = self.load_excel()
        sheet = wb.active
        rows = sheet.max_row
        return rows

    # 获取列数
    def get_cols(self):
        wb = self.load_excel()
        sheet = wb.active
        cols = sheet.max_column
        return cols

    # 获取某一行的数据
    def get_row_value(self, row):
        row_data = []
        wb = self.load_excel()
        sheet = wb.active
        for i in sheet[row]:
            row_data.append(i.value)

        return row_data

    # 获取单元格的内容
    def get_cell_value(self,row , col):
        wb = self.load_excel()
        sheet = wb.active
        data = sheet.cell(row, col).value
        return data

    # 向单元格写入数据
    def excel_write_data(self, row, col, data):
        wb = self.load_excel()
        sheet = wb.active
        sheet.cell(row, col, data)
        wb.save(os.path.abspath(os.path.join(base_path, "..")) + "/Case/mmall testcase.xlsx")


# 单例模式
handleExcel = HandleExcel()

if __name__ == '__main__':
    handleExcel.excel_write_data(1,14,"pass")